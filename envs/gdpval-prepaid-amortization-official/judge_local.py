"""Run the official GDPval rubric through a direct Anthropic Messages API call.

This judge has no external agent-session dependency: it extracts the
candidate workbook's sheets/formulas/values with openpyxl, attaches the
source PDFs as native document blocks, and asks Claude to grade every
official rubric item with a strict binary (pass/fail) decision in a single
request/response — no sandbox, no file uploads, no multi-turn session.
"""

from __future__ import annotations

import base64
import json
import os
import re
import time
from pathlib import Path
from typing import Any

import httpx
import yaml

PROMPT_VERSION = "gdpval_prepaid_official_rubric_judge_v3_direct_api"
DEFAULT_MODEL = "claude-opus-4-5-20251101"
DEFAULT_API_URL = "https://api.anthropic.com/v1/messages"
ANTHROPIC_VERSION = "2023-06-01"


class JudgeConfig:
    def __init__(
        self,
        *,
        api_key: str,
        base_url: str = DEFAULT_API_URL,
        model: str = DEFAULT_MODEL,
        timeout: int = 600,
    ) -> None:
        self.api_key = api_key
        self.base_url = base_url
        self.model = model
        self.timeout = timeout


def load_config() -> JudgeConfig | None:
    config = _load_yaml_config()
    shared = config.get("llm_judge") if isinstance(config.get("llm_judge"), dict) else {}
    section = config.get("gdpval_prepaid_amortization_official")
    section = section if isinstance(section, dict) else {}
    judge = section.get("judge") if isinstance(section.get("judge"), dict) else {}

    api_key = (
        os.environ.get("GDPVAL_JUDGE_API_KEY")
        or os.environ.get("ANTHROPIC_API_KEY")
        or str(judge.get("api_key") or shared.get("api_key") or "")
    )
    if not api_key:
        return None
    return JudgeConfig(
        api_key=api_key,
        base_url=(
            os.environ.get("GDPVAL_JUDGE_BASE_URL")
            or str(judge.get("base_url") or shared.get("base_url") or DEFAULT_API_URL)
        ),
        model=(
            os.environ.get("GDPVAL_JUDGE_MODEL")
            or os.environ.get("LLM_JUDGE_MODEL")
            or str(judge.get("model") or shared.get("model") or DEFAULT_MODEL)
        ),
        timeout=int(
            os.environ.get("GDPVAL_JUDGE_TIMEOUT")
            or os.environ.get("LLM_JUDGE_TIMEOUT")
            or judge.get("timeout")
            or shared.get("timeout")
            or 600
        ),
    )


def run_rubric_judge(
    *,
    candidate_workbooks: list[Path],
    source_files: list[Path],
    expert_workbook: Path,
    rubric_path: Path,
    task_prompt: str,
    artifact_dir: Path,
) -> dict[str, Any]:
    config = load_config()
    if config is None:
        return {
            "ok": False,
            "score_100": 0,
            "error": "judge 未配置；请设置 arena.yaml 的 llm_judge 或 GDPVAL_JUDGE_*/ANTHROPIC_API_KEY 环境变量",
        }
    try:
        return _run(
            config=config,
            candidate_workbooks=candidate_workbooks,
            source_files=source_files,
            expert_workbook=expert_workbook,
            rubric_path=rubric_path,
            task_prompt=task_prompt,
            artifact_dir=artifact_dir,
        )
    except Exception as exc:
        return {"ok": False, "score_100": 0, "error": f"LLM judge failed: {exc}"}


def _run(
    *,
    config: JudgeConfig,
    candidate_workbooks: list[Path],
    source_files: list[Path],
    expert_workbook: Path,
    rubric_path: Path,
    task_prompt: str,
    artifact_dir: Path,
) -> dict[str, Any]:
    rubric = _load_rubric(rubric_path)
    artifact_dir.mkdir(parents=True, exist_ok=True)
    started = time.time()

    candidate_dumps = [
        (path.name, _dump_workbook(path)) for path in candidate_workbooks
    ]
    expert_dump = _dump_workbook(expert_workbook)

    prompt = build_prompt(
        task_prompt=task_prompt,
        rubric=rubric,
        candidate_names=[name for name, _ in candidate_dumps],
    )
    (artifact_dir / "judge_prompt.txt").write_text(prompt, encoding="utf-8")

    content: list[dict[str, Any]] = [{"type": "text", "text": prompt}]
    for name, dump in candidate_dumps:
        content.append({"type": "text", "text": f"\n=== Candidate workbook: {name} ===\n{dump}"})
    content.append({"type": "text", "text": f"\n=== Expert reference workbook: {expert_workbook.name} ===\n{expert_dump}"})
    for source in source_files:
        block = _pdf_block(source)
        if block is not None:
            content.append(block)
        else:
            content.append({"type": "text", "text": f"\n=== Source file (unreadable as PDF): {source.name} ===\n"})

    payload = {
        "model": config.model,
        "max_tokens": 8000,
        "messages": [{"role": "user", "content": content}],
    }
    headers = {
        "x-api-key": config.api_key,
        "anthropic-version": ANTHROPIC_VERSION,
        "content-type": "application/json",
    }
    resp = httpx.post(config.base_url, json=payload, headers=headers, timeout=config.timeout)
    (artifact_dir / "judge_http_status.txt").write_text(str(resp.status_code), encoding="utf-8")
    if resp.status_code >= 400:
        return {"ok": False, "score_100": 0, "error": f"judge API error {resp.status_code}: {resp.text[:500]}"}

    data = resp.json()
    raw_text = "".join(
        block.get("text", "") for block in data.get("content", []) if block.get("type") == "text"
    ).strip()
    (artifact_dir / "judge_raw_response.txt").write_text(raw_text, encoding="utf-8")
    if not raw_text:
        return {"ok": False, "score_100": 0, "error": "judge produced no text response"}

    try:
        review = normalize_result(parse_model_json(raw_text), rubric)
    except Exception as exc:
        return {"ok": False, "score_100": 0, "error": f"judge response parse/validation failed: {exc}"}

    review["model"] = review.get("model") or config.model
    review["prompt_version"] = PROMPT_VERSION
    review["provider"] = "anthropic"
    review["latency_ms"] = int((time.time() - started) * 1000)
    (artifact_dir / "judge_result.json").write_text(
        json.dumps(review, ensure_ascii=False, indent=2), encoding="utf-8"
    )
    return {"ok": True, "score_100": score_100(review), "judge_review": review}


def _dump_workbook(path: Path) -> str:
    """Render an xlsx workbook's sheet names, cell values, and formulas as text."""
    if not path.is_file():
        return f"(missing file: {path.name})"
    try:
        import openpyxl
    except Exception as exc:
        return f"(openpyxl unavailable: {exc})"
    try:
        wb_values = openpyxl.load_workbook(path, data_only=True)
        wb_formulas = openpyxl.load_workbook(path, data_only=False)
    except Exception as exc:
        return f"(failed to open workbook: {exc})"

    lines: list[str] = []
    for sheet_name in wb_values.sheetnames:
        ws_v = wb_values[sheet_name]
        ws_f = wb_formulas[sheet_name]
        lines.append(f"--- Sheet: {sheet_name} ({ws_v.max_row}x{ws_v.max_column}) ---")
        for row_v, row_f in zip(ws_v.iter_rows(), ws_f.iter_rows()):
            cells = []
            for cell_v, cell_f in zip(row_v, row_f):
                value = cell_v.value
                formula = cell_f.value
                if value is None and formula is None:
                    continue
                if isinstance(formula, str) and formula.startswith("="):
                    cells.append(f"{cell_v.coordinate}={formula} -> {value!r}")
                else:
                    cells.append(f"{cell_v.coordinate}={value!r}")
            if cells:
                lines.append(" | ".join(cells))
    return "\n".join(lines) if lines else "(empty workbook)"


def _pdf_block(path: Path) -> dict[str, Any] | None:
    if not path.is_file() or path.suffix.lower() != ".pdf":
        return None
    try:
        data = base64.standard_b64encode(path.read_bytes()).decode("ascii")
    except Exception:
        return None
    return {
        "type": "document",
        "source": {"type": "base64", "media_type": "application/pdf", "data": data},
        "title": path.name,
    }


def build_prompt(*, task_prompt: str, rubric: list[dict[str, Any]], candidate_names: list[str]) -> str:
    compact_rubric = [
        {
            "rubric_item_id": item["rubric_item_id"],
            "max_points": item["score"],
            "criterion": item["criterion"],
        }
        for item in rubric
    ]
    return f"""You are the independent evaluator for an official GDPval accounting task. Grade the Agent deliverable against every rubric item. The candidate workbook content (sheets, cell values, and formulas) and the source PDFs are attached below/after this message as text and document blocks respectively.

Task prompt:
{task_prompt}

Attached materials:
- Candidate deliverable(s): {json.dumps(candidate_names, ensure_ascii=False)} (dumped as text: sheet name / cell coordinate / value / formula)
- Source PDFs: authoritative invoices supplied to the Agent (attached as PDF documents)
- Expert reference workbook: dumped as text, for calibration only — not a pixel-identical answer key

Evaluation procedure:
1. Read the candidate workbook dump carefully: sheet names, formulas, and cell values for all invoice-level rows.
2. Cross-check against the attached source PDFs for criteria involving invoice completeness, amounts, dates, terms, classification, additions, and unsupported adjustments.
3. Use the expert workbook dump as a calibration/reference, while allowing any alternative implementation that satisfies the task and rubric.
4. Grade all {len(rubric)} rubric items independently using STRICT BINARY decisions. Set `passed` to true only when the criterion is fully satisfied; otherwise set it to false. There is no partial credit. A passed item receives its full `max_points`; a failed item receives zero.
5. Reasons must cite concrete workbook sheets/cells/rows or source records where possible. Do not award points merely because the Agent says something is correct.
6. A missing, unreadable, or extra deliverable workbook must be reflected in the relevant rubric items. Do not invent evidence.

Official rubric items:
{json.dumps(compact_rubric, ensure_ascii=False)}

Respond with ONLY the following JSON object (no markdown fences, no other text):
{{
  "schema_version": "gdpval_rubric_judge.v1",
  "overall": {{
    "awarded_points": 0,
    "max_points": {sum(int(item['score']) for item in rubric)},
    "confidence": "high | medium | low",
    "summary": "concise overall assessment"
  }},
  "rubric_scores": [
    {{
      "rubric_item_id": "exact id from the rubric",
      "passed": false,
      "max_points": 0,
      "reason": "specific evidence-based reason"
    }}
  ],
  "unverified_items": ["rubric item ids that could not be reliably inspected"],
  "human_attention_points": ["important ambiguities, if any"]
}}
"""


def normalize_result(result: dict[str, Any], rubric: list[dict[str, Any]]) -> dict[str, Any]:
    scores = result.get("rubric_scores")
    if not isinstance(scores, list):
        raise ValueError("rubric_scores must be a list")
    by_id: dict[str, dict[str, Any]] = {}
    for item in scores:
        if not isinstance(item, dict):
            continue
        item_id = str(item.get("rubric_item_id") or "")
        if item_id in by_id:
            raise ValueError(f"duplicate rubric item: {item_id}")
        by_id[item_id] = item

    normalized: list[dict[str, Any]] = []
    for expected in rubric:
        item_id = str(expected["rubric_item_id"])
        item = by_id.get(item_id)
        if item is None:
            raise ValueError(f"missing rubric item: {item_id}")
        max_points = int(expected["score"])
        passed = item.get("passed")
        if isinstance(passed, bool):
            awarded = max_points if passed else 0
        else:
            try:
                awarded = int(item.get("awarded"))
            except (TypeError, ValueError) as exc:
                raise ValueError(f"missing binary passed decision for {item_id}") from exc
            if awarded not in {0, max_points}:
                raise ValueError(
                    f"non-binary awarded points for {item_id}: {awarded}/{max_points}"
                )
        normalized.append(
            {
                "rubric_item_id": item_id,
                "passed": awarded == max_points,
                "awarded": awarded,
                "max_points": max_points,
                "reason": str(item.get("reason") or ""),
            }
        )

    awarded_total = sum(item["awarded"] for item in normalized)
    max_total = sum(item["max_points"] for item in normalized)
    overall = result.get("overall") if isinstance(result.get("overall"), dict) else {}
    confidence = overall.get("confidence")
    if confidence not in {"high", "medium", "low"}:
        confidence = "medium"
    result["schema_version"] = "gdpval_rubric_judge.v1"
    result["rubric_scores"] = normalized
    result["overall"] = {
        "awarded_points": awarded_total,
        "max_points": max_total,
        "confidence": confidence,
        "summary": str(overall.get("summary") or ""),
    }
    result["unverified_items"] = _string_list(result.get("unverified_items"))
    result["human_attention_points"] = _string_list(result.get("human_attention_points"))
    return result


def score_100(result: dict[str, Any]) -> int:
    overall = result["overall"]
    maximum = int(overall["max_points"])
    return round(100 * int(overall["awarded_points"]) / maximum) if maximum else 0


def parse_model_json(raw: str) -> dict[str, Any]:
    text = raw.strip()
    fenced = re.search(r"```(?:json)?\s*(.*?)```", text, flags=re.DOTALL | re.IGNORECASE)
    if fenced:
        text = fenced.group(1).strip()
    try:
        parsed = json.loads(text)
        if isinstance(parsed, dict) and isinstance(parsed.get("rubric_scores"), list):
            return parsed
    except json.JSONDecodeError:
        pass
    decoder = json.JSONDecoder()
    for match in re.finditer(r"\{", text):
        try:
            parsed, _ = decoder.raw_decode(text[match.start() :])
        except json.JSONDecodeError:
            continue
        if isinstance(parsed, dict) and isinstance(parsed.get("rubric_scores"), list):
            return parsed
    raise ValueError("no complete judge JSON object with rubric_scores found")


def _load_rubric(path: Path) -> list[dict[str, Any]]:
    data = json.loads(path.read_text(encoding="utf-8"))
    rubric = data.get("rubric_json")
    if not isinstance(rubric, list) or not rubric:
        raise ValueError("official rubric_json is missing or empty")
    return rubric


def _repo_root() -> Path:
    return Path(__file__).resolve().parents[2]


def _load_yaml_config() -> dict[str, Any]:
    path = Path(os.environ.get("LANE_CONFIG_PATH") or _repo_root() / "arena.yaml")
    if not path.is_file():
        return {}
    try:
        loaded = yaml.safe_load(path.read_text(encoding="utf-8")) or {}
    except Exception:
        return {}
    return loaded if isinstance(loaded, dict) else {}


def _string_list(value: Any) -> list[str]:
    return [str(item) for item in value] if isinstance(value, list) else []
